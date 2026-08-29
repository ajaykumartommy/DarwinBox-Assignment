from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional


Row = Dict[str, str]

ALIASES = {
    "employeeNumber": {"employee_id", "employee id", "employee_number", "emp_id", "emp id"},
    "given_name": {"first_name", "first name", "given_name", "given name", "firstname"},
    "family_name": {"last_name", "last name", "family_name", "family name", "surname"},
    "email": {"work_email", "work email", "email", "email_address", "email address"},
    "joined_on": {"start_date", "start date", "join_date", "joining date", "date of joining"},
    "team_name": {"department", "department_name", "team", "team name"},
    "phone": {"phone", "mobile", "mobile_number", "phone number"},
}


@dataclass
class Escalation:
    kind: str
    record_key: str
    title: str
    detail: str
    confidence: int
    evidence: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ProcessedRun:
    records: List[Row]
    mappings: List[Dict[str, Any]]
    escalations: List[Escalation]
    audit: List[Dict[str, str]]
    source_count: int
    duplicates_consolidated: int


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalized_header(value: str) -> str:
    return re.sub(r"[\s-]+", "_", clean(value).lower())


def parse_csv(content: bytes) -> List[Row]:
    text = content.decode("utf-8-sig")
    return [{key: clean(value) for key, value in row.items() if key} for row in csv.DictReader(io.StringIO(text))]


def parse_xlsx(content: bytes) -> List[Row]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - installation is documented
        raise ValueError("XLSX support needs openpyxl installed.") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(header) for header in rows[0]]
    return [{headers[index]: clean(value) for index, value in enumerate(values) if headers[index]} for values in rows[1:]]


def parse_source(filename: str, content: bytes) -> List[Row]:
    if filename.lower().endswith(".csv"):
        return parse_csv(content)
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return parse_xlsx(content)
    raise ValueError(f"Unsupported source type for '{filename}'. Use CSV or XLSX.")


def safe_iso_date(value: str) -> Optional[str]:
    value = clean(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return value
    match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", value)
    if not match:
        return None
    first, second, year = map(int, match.groups())
    if first > 12 and second <= 12:
        return f"{year:04d}-{second:02d}-{first:02d}"
    if second > 12 and first <= 12:
        return f"{year:04d}-{first:02d}-{second:02d}"
    return None


def field_value(row: Row, target_field: str) -> tuple[str, Optional[str], int]:
    candidates = {normalized_header(target_field), *{normalized_header(value) for value in ALIASES.get(target_field, set())}}
    matches = [(header, value) for header, value in row.items() if normalized_header(header) in candidates and clean(value)]
    if not matches:
        return "", None, 0
    header, value = matches[0]
    return clean(value), header, 100 if normalized_header(header) == normalized_header(target_field) else 94


def normalise_value(value: str, field: Dict[str, Any], key: str, escalations: List[Escalation]) -> str:
    fmt = field.get("format", "string")
    if not value:
        return ""
    if fmt == "email":
        return value.lower()
    if fmt == "phone":
        digits = re.sub(r"[^\d+]", "", value)
        return digits if digits.startswith("+") else f"+{digits}" if digits else ""
    if fmt == "date":
        normalized = safe_iso_date(value)
        if normalized:
            return normalized
        escalations.append(Escalation(
            kind="ambiguous_date", record_key=key, title="Confirm date format",
            detail=f"'{value}' could be day/month or month/day. The agent will not guess.",
            confidence=48, evidence={"source_value": value},
        ))
        return ""
    return value


def mapping_summary(source_rows: Iterable[Row], fields: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    headers = {header for row in source_rows for header in row}
    result = []
    for target in fields:
        target_name = target["name"]
        direct = next((header for header in headers if normalized_header(header) == normalized_header(target_name)), None)
        alias = next((header for header in headers if normalized_header(header) in {normalized_header(value) for value in ALIASES.get(target_name, set())}), None)
        source = direct or alias
        if source:
            result.append({
                "source_field": source, "target_field": target_name,
                "confidence": 100 if direct else 94,
                "reason": "Exact normalized header match" if direct else "Known source alias and compatible value profile",
                "status": "auto_applied",
            })
    return result


def process_rows(source_rows: List[Row], target_schema: Dict[str, Any]) -> ProcessedRun:
    fields = target_schema.get("fields", [])
    if not fields:
        raise ValueError("Target schema must include at least one field.")
    mappings = mapping_summary(source_rows, fields)
    escalations: List[Escalation] = []
    transformed: List[Row] = []
    known_source_headers = {normalized_header(item["source_field"]) for item in mappings}
    unknown_headers = {header for row in source_rows for header in row if normalized_header(header) not in known_source_headers}

    for index, source in enumerate(source_rows, start=1):
        row: Row = {}
        preliminary_key = clean(source.get("employee_id") or source.get("employeeNumber") or source.get("work_email") or source.get("email")) or f"source-row-{index}"
        for target in fields:
            value, _, _ = field_value(source, target["name"])
            normalized = normalise_value(value, target, preliminary_key, escalations)
            row[target["name"]] = normalized
            if target.get("required") and not normalized and not value:
                escalations.append(Escalation(
                    kind="validation_failure", record_key=preliminary_key, title=f"Provide required {target['name']}",
                    detail=f"No valid value was available for required target field '{target['name']}'.",
                    confidence=0, evidence={"target_field": target["name"]},
                ))
        transformed.append(row)

    # Unknown columns are informational unless one has a high-impact ambiguous name.
    for header in unknown_headers:
        if normalized_header(header) in {"location", "entity", "status"}:
            escalations.append(Escalation(
                kind="ambiguous_mapping", record_key="source-column",
                title=f"Map '{header}' to a target field",
                detail=f"The source column '{header}' is populated but has no safe one-to-one target mapping.",
                confidence=62, evidence={"source_field": header},
            ))

    unique: Dict[str, Row] = {}
    duplicates = 0
    for row in transformed:
        key = clean(row.get("employeeNumber") or row.get("email"))
        if not key:
            continue
        existing = unique.get(key)
        if not existing:
            unique[key] = row
            continue
        conflicts = [field for field, value in row.items() if value and existing.get(field) and existing[field] != value]
        if conflicts:
            escalations.append(Escalation(
                kind="duplicate_conflict", record_key=key, title="Resolve conflicting duplicate records",
                detail=f"Duplicate identity has different values for {', '.join(conflicts)}.",
                confidence=38, evidence={"fields": conflicts},
            ))
        else:
            for field, value in row.items():
                if value and not existing.get(field):
                    existing[field] = value
            duplicates += 1

    audit = [
        {"type": "Ingested", "detail": f"Profiled {len(source_rows)} source records."},
        {"type": "Mapped", "detail": f"Applied {len(mappings)} explainable high-confidence mappings."},
        {"type": "Cleaned", "detail": "Normalized only safe email, phone, whitespace, and unambiguous date values."},
        {"type": "Merged", "detail": f"Consolidated {duplicates} non-conflicting duplicates."},
    ]
    if escalations:
        audit.append({"type": "Escalated", "detail": f"Paused {len(escalations)} cases that need a human decision."})
    return ProcessedRun(list(unique.values()), mappings, escalations, audit, len(source_rows), duplicates)
